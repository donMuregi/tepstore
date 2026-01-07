from django.contrib import admin
from django.utils.html import format_html
from django.shortcuts import render, redirect
from django.urls import path
from django import forms
from django.contrib import messages
import openpyxl
from io import BytesIO
from .models import (
    Category, Product, ProductImage, ProductVariant, Review, Brand,
    FinancingPlan, FinancingApplication,
    EnterpriseBundle, EnterpriseOrder,
    EducationBoard, ClassroomPackage, Fundraiser, DonationAmount, Donation,
    EducationTablet, TabletSoftware, SchoolTabletOrder, SchoolTabletOrderItem,
    Cart, CartItem, Order, OrderItem, HeroSlide, TradeInRequest, Employer, Bank, School, Policy
)


# ============ CUSTOMIZE ADMIN SITE ============

admin.site.site_header = "TEP Digital Admin"
admin.site.site_title = "TEP Digital"
admin.site.index_title = "Store Management"


# ============================================================================
#                           SITE SETTINGS
# ============================================================================

@admin.register(Policy)
class PolicyAdmin(admin.ModelAdmin):
    list_display = ['policy_type', 'title', 'last_updated', 'is_active', 'updated_at']
    list_filter = ['policy_type', 'is_active']
    list_editable = ['is_active']
    search_fields = ['title', 'content']
    ordering = ['policy_type']
    
    fieldsets = (
        ('Policy Information', {
            'fields': ('policy_type', 'title', 'last_updated')
        }),
        ('Content', {
            'fields': ('content',),
            'description': 'You can use HTML tags for formatting'
        }),
        ('Settings', {
            'fields': ('is_active',)
        }),
    )


@admin.register(HeroSlide)
class HeroSlideAdmin(admin.ModelAdmin):
    list_display = ['title', 'highlight', 'image_preview', 'order', 'is_active', 'updated_at']
    list_filter = ['is_active', 'card_icon']
    list_editable = ['order', 'is_active']
    search_fields = ['title', 'highlight', 'description']
    ordering = ['order']
    
    fieldsets = (
        ('Content', {
            'fields': ('badge_text', 'title', 'highlight', 'description')
        }),
        ('Buttons', {
            'fields': (
                ('primary_button_text', 'primary_button_link'),
                ('secondary_button_text', 'secondary_button_link'),
            )
        }),
        ('Image', {
            'fields': ('image',)
        }),
        ('Price Card', {
            'fields': ('card_icon', 'card_label', 'card_price', 'card_subtext')
        }),
        ('Appearance', {
            'fields': ('background_color',),
        }),
        ('Settings', {
            'fields': ('order', 'is_active')
        }),
    )
    
    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" width="80" height="50" style="object-fit: cover; border-radius: 4px;" />', obj.image.url)
        return "No image"
    image_preview.short_description = 'Preview'


# ============================================================================
#                           FINANCING PARTNERS
# ============================================================================

class ExcelUploadForm(forms.Form):
    excel_file = forms.FileField(label='Excel File', help_text='Upload an Excel file with columns: Name, Code, Address, Contact Person, Contact Email, Contact Phone')


@admin.register(Employer)
class EmployerAdmin(admin.ModelAdmin):
    list_display = ['name', 'code', 'contact_person', 'contact_email', 'is_active', 'created_at']
    list_filter = ['is_active']
    list_editable = ['is_active']
    search_fields = ['name', 'code', 'contact_person']
    ordering = ['name']
    
    change_list_template = 'admin/store/employer/change_list.html'
    
    fieldsets = (
        ('Employer Information', {
            'fields': ('name', 'code', 'address')
        }),
        ('Contact Details', {
            'fields': ('contact_person', 'contact_email', 'contact_phone')
        }),
        ('Settings', {
            'fields': ('is_active',)
        }),
    )
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel), name='store_employer_upload_excel'),
        ]
        return custom_urls + urls
    
    def upload_excel(self, request):
        if request.method == 'POST':
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    wb = openpyxl.load_workbook(BytesIO(excel_file.read()))
                    ws = wb.active
                    
                    created_count = 0
                    updated_count = 0
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            name = str(row[0]).strip()
                            code = str(row[1]).strip() if row[1] else None
                            address = str(row[2]).strip() if row[2] else ''
                            contact_person = str(row[3]).strip() if row[3] else ''
                            contact_email = str(row[4]).strip() if row[4] else ''
                            contact_phone = str(row[5]).strip() if row[5] else ''
                            
                            employer, created = Employer.objects.update_or_create(
                                name=name,
                                defaults={
                                    'code': code,
                                    'address': address,
                                    'contact_person': contact_person,
                                    'contact_email': contact_email,
                                    'contact_phone': contact_phone,
                                    'is_active': True,
                                }
                            )
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                    
                    messages.success(request, f'Imported {created_count} new, updated {updated_count} existing employers.')
                    return redirect('..')
                except Exception as e:
                    messages.error(request, f'Error processing file: {str(e)}')
        else:
            form = ExcelUploadForm()
        
        context = {
            'form': form,
            'title': 'Upload Employers from Excel',
            'opts': self.model._meta,
        }
        return render(request, 'admin/store/employer/upload_excel.html', context)


@admin.register(Bank)
class BankAdmin(admin.ModelAdmin):
    list_display = ['name', 'code', 'logo_preview', 'contact_email', 'is_active', 'created_at']
    list_filter = ['is_active']
    list_editable = ['is_active']
    search_fields = ['name', 'code']
    ordering = ['name']
    
    fieldsets = (
        ('Bank Information', {
            'fields': ('name', 'code', 'logo', 'branch')
        }),
        ('Contact Details', {
            'fields': ('contact_email', 'contact_phone')
        }),
        ('Integration', {
            'fields': ('api_endpoint',),
            'description': 'API endpoint for bank credit check integration'
        }),
        ('Settings', {
            'fields': ('is_active',)
        }),
    )
    
    def logo_preview(self, obj):
        if obj.logo:
            return format_html('<img src="{}" width="40" height="40" style="object-fit: contain;" />', obj.logo.url)
        return "-"
    logo_preview.short_description = 'Logo'


# ============================================================================
#                           PRODUCTS & CATALOG
# ============================================================================

class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1


class ProductVariantInline(admin.TabularInline):
    model = ProductVariant
    extra = 1


@admin.register(Brand)
class BrandAdmin(admin.ModelAdmin):
    list_display = ['name', 'slug']
    prepopulated_fields = {'slug': ('name',)}


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'slug', 'category_type', 'created_at']
    list_filter = ['category_type']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name']


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ['name', 'brand', 'category', 'product_type', 'price', 'sale_price', 'stock', 'is_active', 'is_featured']
    list_filter = ['is_active', 'is_featured', 'category', 'product_type', 'brand']
    list_editable = ['price', 'sale_price', 'stock', 'is_active', 'is_featured']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name', 'description']
    inlines = [ProductImageInline, ProductVariantInline]


@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    list_display = ['product', 'user', 'rating', 'created_at']
    list_filter = ['rating', 'created_at']


# ============================================================================
#                           BNPL / FINANCING
# ============================================================================

@admin.register(FinancingPlan)
class FinancingPlanAdmin(admin.ModelAdmin):
    list_display = ['months', 'interest_rate', 'is_active']
    list_editable = ['interest_rate', 'is_active']


@admin.register(FinancingApplication)
class FinancingApplicationAdmin(admin.ModelAdmin):
    list_display = ['application_id', 'full_name', 'application_type', 'employer_display', 'bank_display', 'product', 'status', 'created_at']
    list_filter = ['application_type', 'status', 'financing_plan', 'employer', 'bank']
    search_fields = ['full_name', 'id_number', 'employer_name', 'organization_name']
    readonly_fields = ['application_id']
    
    fieldsets = (
        ('Application Info', {
            'fields': ('application_id', 'application_type', 'status')
        }),
        ('Applicant Details', {
            'fields': ('full_name', 'id_number', 'kra_pin')
        }),
        ('Employer / Bank Selection', {
            'fields': ('employer', 'employer_name', 'staff_number', 'bank'),
            'description': 'Individual: Select employer, or bank if not listed. Chama/Corporate: Select bank.'
        }),
        ('Organization Details (Chama/Corporate)', {
            'fields': ('organization_name', 'registration_number', 'certificate'),
            'classes': ('collapse',)
        }),
        ('Product & Financing', {
            'fields': ('product', 'variant', 'financing_plan')
        }),
        ('Approval Details', {
            'fields': ('approved_amount', 'monthly_payment', 'bank_response')
        }),
    )
    
    def employer_display(self, obj):
        if obj.employer:
            return obj.employer.name
        return obj.employer_name or '-'
    employer_display.short_description = 'Employer'
    
    def bank_display(self, obj):
        if obj.bank:
            return obj.bank.name
        return obj.preferred_bank or '-'
    bank_display.short_description = 'Bank'


# ============================================================================
#                           TELCO CONTRACTS / ENTERPRISE
# ============================================================================

@admin.register(EnterpriseBundle)
class EnterpriseBundleAdmin(admin.ModelAdmin):
    list_display = ['name', 'product', 'data_gb', 'minutes', 'minimum_quantity', 'price_per_device', 'is_active']
    list_filter = ['is_active']
    list_editable = ['price_per_device', 'is_active']


@admin.register(EnterpriseOrder)
class EnterpriseOrderAdmin(admin.ModelAdmin):
    list_display = ['order_id', 'company_name', 'bundle', 'quantity', 'total_amount', 'status', 'created_at']
    list_filter = ['status', 'preferred_bank']
    search_fields = ['company_name', 'contact_person', 'contact_email']
    readonly_fields = ['order_id']


# ============================================================================
#                           EDUCATION
# ============================================================================

class SchoolExcelUploadForm(forms.Form):
    excel_file = forms.FileField(label='Excel File', help_text='Upload Excel with: Name, County, Location, School Type')


@admin.register(School)
class SchoolAdmin(admin.ModelAdmin):
    list_display = ['name', 'location', 'county', 'school_type', 'is_approved', 'created_at']
    list_filter = ['is_approved', 'county', 'school_type']
    list_editable = ['is_approved']
    search_fields = ['name', 'location', 'county']
    ordering = ['name']
    
    change_list_template = 'admin/store/school/change_list.html'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel), name='store_school_upload_excel'),
        ]
        return custom_urls + urls
    
    def upload_excel(self, request):
        if request.method == 'POST':
            form = SchoolExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    wb = openpyxl.load_workbook(BytesIO(excel_file.read()))
                    created_count = 0
                    updated_count = 0
                    
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0]:
                                name = str(row[0]).strip()
                                county = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                                location = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                                school_type = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                                
                                school, created = School.objects.update_or_create(
                                    name=name,
                                    defaults={
                                        'location': location,
                                        'county': county,
                                        'school_type': school_type,
                                        'is_approved': True,
                                    }
                                )
                                if created:
                                    created_count += 1
                                else:
                                    updated_count += 1
                    
                    messages.success(request, f'Imported {created_count} new, updated {updated_count} existing schools.')
                    return redirect('..')
                except Exception as e:
                    messages.error(request, f'Error processing file: {str(e)}')
        else:
            form = SchoolExcelUploadForm()
        
        context = {
            'form': form,
            'title': 'Upload Schools from Excel',
            'opts': self.model._meta,
        }
        return render(request, 'admin/store/school/upload_excel.html', context)


@admin.register(EducationBoard)
class EducationBoardAdmin(admin.ModelAdmin):
    list_display = ['name', 'price', 'installation_included', 'is_active']
    prepopulated_fields = {'slug': ('name',)}
    list_editable = ['price', 'is_active']


@admin.register(ClassroomPackage)
class ClassroomPackageAdmin(admin.ModelAdmin):
    list_display = ['name', 'boards_included', 'price', 'installation_included', 'is_active']
    prepopulated_fields = {'slug': ('name',)}
    list_editable = ['price', 'is_active']


@admin.register(DonationAmount)
class DonationAmountAdmin(admin.ModelAdmin):
    list_display = ['amount_usd', 'is_active']
    list_editable = ['is_active']


class DonationInline(admin.TabularInline):
    model = Donation
    extra = 0
    readonly_fields = ['donation_id', 'donor_name', 'amount', 'status', 'created_at']


@admin.register(Fundraiser)
class FundraiserAdmin(admin.ModelAdmin):
    list_display = ['fundraiser_id', 'school_name', 'creator', 'fundraiser_type', 'target_amount', 'current_amount', 'status']
    list_filter = ['fundraiser_type', 'status']
    search_fields = ['school_name', 'creator__username']
    readonly_fields = ['fundraiser_id', 'share_link', 'current_amount']
    inlines = [DonationInline]


@admin.register(Donation)
class DonationAdmin(admin.ModelAdmin):
    list_display = ['donation_id', 'fundraiser', 'donor_name', 'amount', 'payment_method', 'status', 'created_at']
    list_filter = ['status', 'payment_method']
    search_fields = ['donor_name', 'donor_email']
    readonly_fields = ['donation_id']


@admin.register(EducationTablet)
class EducationTabletAdmin(admin.ModelAdmin):
    list_display = ['name', 'brand', 'size', 'price', 'stock', 'is_active']
    list_filter = ['brand', 'size', 'is_active']
    prepopulated_fields = {'slug': ('name',)}
    list_editable = ['price', 'stock', 'is_active']


@admin.register(TabletSoftware)
class TabletSoftwareAdmin(admin.ModelAdmin):
    list_display = ['name', 'price', 'is_default']
    prepopulated_fields = {'slug': ('name',)}
    list_editable = ['price', 'is_default']


class SchoolTabletOrderItemInline(admin.TabularInline):
    model = SchoolTabletOrderItem
    extra = 0


@admin.register(SchoolTabletOrder)
class SchoolTabletOrderAdmin(admin.ModelAdmin):
    list_display = ['order_id', 'school_name', 'status', 'total_amount', 'created_at']
    list_filter = ['status']
    search_fields = ['school_name', 'school_email']
    readonly_fields = ['order_id']
    inlines = [SchoolTabletOrderItemInline]


# ============================================================================
#                           ORDERS & CART
# ============================================================================

class CartItemInline(admin.TabularInline):
    model = CartItem
    extra = 0


@admin.register(Cart)
class CartAdmin(admin.ModelAdmin):
    list_display = ['cart_id', 'user', 'item_count', 'total', 'created_at']
    readonly_fields = ['cart_id']
    inlines = [CartItemInline]


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ['product', 'variant', 'quantity', 'unit_price']


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ['order_id', 'full_name', 'email', 'total', 'status', 'payment_status', 'created_at']
    list_filter = ['status', 'payment_status']
    search_fields = ['full_name', 'email', 'phone', 'order_id']
    readonly_fields = ['order_id', 'subtotal', 'total']
    inlines = [OrderItemInline]


# ============================================================================
#                           TRADE-IN REQUESTS
# ============================================================================

@admin.register(TradeInRequest)
class TradeInRequestAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'phone', 'current_device', 'device_condition', 'product_name', 'status', 'created_at']
    list_filter = ['status', 'device_condition', 'created_at']
    search_fields = ['name', 'email', 'phone', 'current_device', 'product_name']
    list_editable = ['status']
    readonly_fields = ['created_at', 'updated_at']
    
    fieldsets = (
        ('Customer Information', {
            'fields': ('name', 'email', 'phone')
        }),
        ('Trade-In Device', {
            'fields': ('current_device', 'device_condition', 'message')
        }),
        ('Product Requested', {
            'fields': ('product', 'product_name', 'variant', 'variant_name')
        }),
        ('Admin', {
            'fields': ('status', 'estimated_value', 'admin_notes')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
